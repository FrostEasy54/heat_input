import os
import sys
import openpyxl
from PyQt6.QtWidgets import QMainWindow, QHeaderView, QApplication
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from PyQt6 import uic
from rooms import RoomsTable
from people import PeopleTable
from equipment import EquipmentTable
from lamps import LampsTable

ui_path = os.path.abspath(os.path.join(os.path.dirname(__file__), 'mygui.ui'))


class MyGUI(QMainWindow, RoomsTable, PeopleTable, EquipmentTable, LampsTable):

    def __init__(self):
        super(MyGUI, self).__init__()
        uic.loadUi(ui_path, self)

        self.setWindowTitle("Расчёт теплопоступлений в помещениях")
        self.actionAuthors.triggered.connect(self.AuthorsMessage)
        self.actionTutorial.triggered.connect(self.TutorialMessage)
        self.actionNormativeDocs.triggered.connect(self.NormativeDocsMessage)
        # Расстягивание ширины заголовков таблиц под ширину экрана
        self.RoomsTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.PeopleTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.EquipmentTableWidget.horizontalHeader(
        ).setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.LampsTableWidget.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        # Установка списка городов и температуры на лист Помещения
        self.GetClimateCity()
        self.UpdateTempLabel(0)
        # Установка значений движения скорости движения воздуха в помещении
        self.SetWindSpeed()
        self.SeasonComboBox.currentTextChanged.connect(self.SetWindSpeed)

        # Соединение кнопок очистки
        self.actionClearRooms.triggered.connect(self.ClearRoomsTable)
        self.actionClearPeople.triggered.connect(self.ClearPeopleTable)
        self.actionClearEquipment.triggered.connect(self.ClearEquipmentTable)
        self.actionClearLamps.triggered.connect(self.ClearLampsTable)
        self.actionClearAll.triggered.connect(self.ClearAllTables)

        # Соединение кнопок сохранения
        self.ActionSaveRooms.triggered.connect(
            lambda: self.ExportSingleTable(self.ExportRoomsTable))
        self.ActionSavePeople.triggered.connect(
            lambda: self.ExportSingleTable(self.ExportPeopleTable))
        self.ActionSaveEquipment.triggered.connect(
            lambda: self.ExportSingleTable(self.ExportEquipmentTable))
        self.ActionSaveLamps.triggered.connect(lambda: self.ExportSingleTable(
            self.ExportLampsTable))
        self.ActionSaveAll.triggered.connect(self.ExportAllTables)

        # Соединение кнопок импорта
        self.ActionImportRooms.triggered.connect(
            lambda: self.ImportSingleTable(self.ImportRoomsTable))
        self.ActionImportPeople.triggered.connect(
            lambda: self.ImportSingleTable(self.ImportPeopleTable))
        self.ActionImportEquipment.triggered.connect(
            lambda: self.ImportSingleTable(self.ImportEquipmentTable))
        self.ActionImportLamps.triggered.connect(
            lambda: self.ImportSingleTable(self.ImportLampsTable))
        self.ActionImportAll.triggered.connect(self.ImportAllTables)

        # Добавление и удаление новых рядов для таблицы Помещения
        self.AddRowRoomsPushButton.clicked.connect(self.AddRoomsRow)
        self.RemoveRowRoomsPushButton.clicked.connect(self.RemoveRoomsRow)
        # Установка виджетов для таблицы Помещения
        self.RoomTypeComboBox()
        self.RoomNumberSpinBox()
        self.RoomAreaDoubleSpinBox()
        self.SetRoomInnerTemp()
        self.CityComboBox.currentIndexChanged.connect(self.SetRoomInnerTemp)
        self.MakeRoomCellReadOnly()
        self.MakeRoomTempCellReadOnly()
        # Добавление теплопоступлений с других листов на лист Помещения
        self.HeatInputPushButton.clicked.connect(self.AddPeopleHeatInput)
        self.HeatInputPushButton.clicked.connect(self.AddEquipmentHeatInput)
        self.HeatInputPushButton.clicked.connect(self.AddLampsHeatInput)
        # Сумма всех теплопоступлений
        self.HeatInputPushButton.clicked.connect(self.SumOfHeatInput)

        # Добавление и удаление новых рядов для таблицы Люди
        self.AddRowPeoplePushButton.clicked.connect(self.AddPeopleRow)
        self.RemoveRowPeoplePushButton.clicked.connect(self.RemovePeopleRow)
        # Установка виджетов для таблицы Люди
        self.PeopleSexComboBox()
        self.PeopleRoomNumberSpinBox()
        self.PeopleWorkComboBox()
        self.PeopleClothesComboBox()
        self.PeopleCountSpinBox()
        self.MakePeopleCellReadOnly()
        # Расчёт теплопоступлений от людей
        self.PeopleHeatInputPushButton.clicked.connect(self.PeopleHeatInput)

        # Добавление и удаление новых рядов для таблицы Оборудование
        self.AddRowEquipmentPushButton.clicked.connect(self.AddEquipmentRow)
        self.RemoveRowEquipmentPushButton.clicked.connect(
            self.RemoveEquipmentRow)
        # Установка виджетов для таблицы Оборудование
        self.EquipmentRoomNumberSpinBox()
        self.EquipmentTypeComboBox()
        self.EquipmentCountSpinBox()
        self.EquipmentPowerSpinBox()
        self.EquipmentLoadDoubleSpinBox()
        self.MakeEquipmentCellReadOnly()
        self.SetPowerForEquipmentType()
        self.CheckForSetPower()
        # Расчёт теплопоступлений от Оборудования на листе Оборудование
        self.EquipmentHeatInputPushButton.clicked.connect(
            self.EquipmentHeatInput)

        # Добавление и удаление новых рядов для таблицы Светильники
        self.AddRowLampsPushButton.clicked.connect(self.AddLampsRow)
        self.RemoveRowLampsPushButton.clicked.connect(self.RemoveLampsRow)
        # Установка виджетов для таблицы Светильники
        self.LampsRoomNumberSpinBox()
        self.LampsTypeComboBox()
        self.LampsPurposeComboBox()
        self.LampsCountSpinBox()
        self.MakeLampsCellReadOnly()
        # Расчёт теплопоступлений от Светильников на листе Светильники
        self.LampsHeatInputPushButton.clicked.connect(self.LampsHeatInput)

        self.actionHeatInputGraph.triggered.connect(self.showHistogram)

    # Получение списка городов из excel файла
    def GetClimateCity(self):
        path = 'climate_db.xlsx'
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        row = 651
        self.city_temp_db = {}
        for city, temp in zip(ws['A'][9:row], ws['L'][9:row]):
            self.city_temp_db[city.value] = temp.value
        self.CityComboBox.addItems(self.city_temp_db.keys())
        self.CityComboBox.currentIndexChanged.connect(self.UpdateTempLabel)

    # Обновление температуры в зависимости от города
    def UpdateTempLabel(self, index):
        current_city = self.CityComboBox.currentText()
        temp = self.city_temp_db.get(current_city, 'Н/Д')
        self.TempOutsideValueLabel.setText(f'{temp}°C')

    # Установка максимума и минимума скорости движения воздуха
    def SetWindSpeed(self):
        if self.SeasonComboBox.currentText() == "Теплый":
            self.WindSpeedDoubleSpinBox.setMinimum(0.1)
            self.WindSpeedDoubleSpinBox.setMaximum(0.6)
        else:
            self.WindSpeedDoubleSpinBox.setMinimum(0.0)
            self.WindSpeedDoubleSpinBox.setMaximum(0.5)
        self.WindSpeedDoubleSpinBox.setSingleStep(0.1)

    # Очистка всех таблиц
    def ClearAllTables(self):
        self.ClearPeopleTable()
        self.ClearRoomsTable()
        self.ClearEquipmentTable()
        self.ClearLampsTable()

    def ExportSingleTable(self, export_function):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getSaveFileName(
            self, "Сохранить файл", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        wb = openpyxl.Workbook()
        export_function(wb)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        wb.save(path)
        QMessageBox.information(
            self, "Успешно", "Таблица успешно сохранена в файл: {}".format(path))

    # Экспорт всех таблиц в  один Excel файл
    def ExportAllTables(self):
        file_dialog = QFileDialog()
        path, _ = file_dialog.getSaveFileName(
            self, "Сохранить файл", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        wb = openpyxl.Workbook()
        self.ExportRoomsTable(wb)
        self.ExportPeopleTable(wb)
        self.ExportEquipmentTable(wb)
        self.ExportLampsTable(wb)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        wb.save(path)
        QMessageBox.information(
            self, "Успешно", "Все таблицы успешно сохранены в файл: {}".format(path))

    def ImportSingleTable(self, import_function):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx);;All Files (*)")
        path, _ = file_dialog.getOpenFileName(
            self, "Открыть файл", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path)
            import_function(wb)
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Произошла ошибка при открытии файла: {str(e)}")

    def ImportAllTables(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.FileMode.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx);;All Files (*)")
        path, _ = file_dialog.getOpenFileName(
            self, "Открыть файл", "", "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path)
            self.ImportLampsTable(wb)
            self.ImportEquipmentTable(wb)
            self.ImportPeopleTable(wb)
            self.ImportRoomsTable(wb)
        except Exception as e:
            QMessageBox.critical(
                self, "Ошибка", f"Произошла ошибка при открытии файла: {str(e)}")

    def AuthorsMessage(self):
        message_box = QMessageBox()
        message_box.information(
            None, "Авторы", "Данное приложение было разработано студентами НИУ МГСУ 4 курса направления ИСТАС, \nгруппы ИЦТМС 4-2, \nДубровиным В.А. и Мышкиным А.В. \n2023 г.")
        message_box.setFixedSize(500, 200)

    def TutorialMessage(self):
        message_box = QMessageBox()
        message_box.information(None, "Руководство пользователя", "1. Заполнение данных начинается с выбора города, периода года и скорости ветра в помещении. \n2. На листе Люди необходимо заполнить данные по столбцам и при необходимости добавить новые строки. Заметьте, что люди указанные в одном и том же номере помещения будут суммироваться в итоговой таблице. Для рассчёта теплопоступлений нажимите кнопку Рассчиать теплопоступления внизу экрана.\n3. Перейдите на лист Оборудование и заполните его по столбцам и затем добавьте новые строки по необходимости. Рассчет производится так же по нажатию кнопки в нижней части экрана.\n4. Заполните лист Светильники по аналогии с предыдущими листами. \n5. Перейдите на лист Помещения, укажите его номер, название, тип и площадь. Нажмите кнопку Рассчитать теплопоступления для рассчёта итоговых теплопоступлений в помещении. \nПриложение так же поддерживает функции экспорта и импорта данных. Для этого необходимо выбрать соответствующий вашим задачам пункт из верхнего меню.\nСтоит отметить, что импорт данных происходит строго по типам данных, представленных в таблицах в приложении. То есть, номер помещения в вашем Excel файле должен быть числом, название помещения — текстом, площадь помещения — целочисленным или с плавающей точкой и так далее")
        message_box.setFixedSize(1000, 800)

    def NormativeDocsMessage(self):
        message_box = QMessageBox()
        message_box.information(None, "Используемые нормативные документы",
                                "Нормативные документы на основе которых проводились вычисления теплопоступлений от людей, электроприборов и освещения в производственных помещениях:Глава 8.3 - Внутренние санитарно-технические устройства Ч. 1. Отопление, М.: Стройиздат, 1990.\nФормула для расчёта:Q_чел=β_и*β_од*(2,5+10,36 √(υ_в ))*(35-t_п).\nГде:\nβи – коэффициент учета интенсивности работы, равный 1,0 для легкой работы, 1,07 для работы средней тяжести и 1,15 для тяжелой работы;\nβод – коэффициент учета теплозащитных свойств одежды, равный 1,0 для легкой одежды, 0,65 для обычной одежды и 0,40 для утепленной одежды;\nυв – скорость движения воздуха в помещении, м/с;\ntп – температура, °С.\n\nВ бытовых помещениях:п. 3.1 СНиП 2.04.05—86.\nФормула для расчёта: Q_быт=q_1*A_п\nГде:\nq1 – это теплопоступления на 1 м2 площади пола, Вт/м2 [ккал/(ч-м2)]; принимают по данным главы СНиП 2.04.05-86; (В соответствии с п. 3.1 СНиП 2.04.05—86 бытовые тепловыделения Qбыт следует учитывать для жилых комнат и кухонь в размере 21 Вт на 1 м2 площади пола.\nАп – это площадь пола жилой комнаты или кухни, м2.\n\nПри искусственном освещении и работающем электрическом производственном оборудовании тепловыделения Qэ, Вт (ккал/ч), составляют:Q_э=kN\nГде:\nк – общий коэффициент загрузки оборудования (к = 0,5 — 0,8); при светильниках в помещении к — 1,0, светильниках, встроенных, в перекрытие помещения, к = 0,40;\nN – мощность осветительных приборов или силового оборудования, Вт.")


def main():
    app = QApplication(sys.argv)
    window = MyGUI()
    window.show()
    app.exec()


if __name__ == '__main__':
    main()
